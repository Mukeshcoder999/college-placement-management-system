from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Q
from accounts.permissions import IsStudent, IsCompany, IsPlacementOfficer
from students.models import StudentProfile
from students.serializers import StudentProfileSerializer
from companies.models import CompanyProfile
from companies.serializers import CompanyProfileSerializer
from jobs.models import Job
from jobs.serializers import JobSerializer
from applications.models import JobApplication
from applications.serializers import JobApplicationSerializer
from notifications.models import Notification
from django.db.models.functions import TruncMonth
#excel reports

from openpyxl import Workbook
from django.http import HttpResponse

#PDF Exports 
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.units import inch

# Create your views here.

class StudentDashboardAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsStudent,
    ]

    def get(self, request):

        student = request.user.studentprofile
        application_stats = JobApplication.objects.filter(
        student=student
                ).aggregate(
                    total_applications=Count("id"),
                    applied=Count("id", filter=Q(status="Applied")),
                    shortlisted=Count("id", filter=Q(status="Shortlisted")),
                    selected=Count("id", filter=Q(status="Selected")),
                    rejected=Count("id", filter=Q(status="Rejected")),
                )
        

        unread_notifications = Notification.objects.filter(
            recipient=request.user,
            is_read=False
        ).count()

        data = {
            "total_applications": application_stats["total_applications"],
            "applied": application_stats["applied"],
            "shortlisted": application_stats["shortlisted"],
            "selected": application_stats["selected"],
            "rejected": application_stats["rejected"],
            "unread_notifications": unread_notifications,
        }

        return Response(data)

#comapny dashboard api view

class CompanyDashboardAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsCompany,
    ]

    def get(self, request):

        company = request.user.companyprofile

        job_stats = Job.objects.filter(
            company=company
        ).aggregate(

            jobs_posted=Count("id"),

            active_jobs=Count(
                "id",
                filter=Q(is_active=True)
            ),

            inactive_jobs=Count(
                "id",
                filter=Q(is_active=False)
            ),
        )

        application_stats = JobApplication.objects.filter(
            job__company=company
        ).aggregate(

            applications_received=Count("id"),

            shortlisted_students=Count(
                "id",
                filter=Q(status="Shortlisted")
            ),

            selected_students=Count(
                "id",
                filter=Q(status="Selected")
            ),
        )

        unread_notifications = Notification.objects.filter(
            recipient=request.user,
            is_read=False
        ).count()

        data = {
            "jobs_posted": job_stats["jobs_posted"],
            "active_jobs": job_stats["active_jobs"],
            "inactive_jobs": job_stats["inactive_jobs"],
            "applications_received": application_stats["applications_received"],
            "shortlisted_students": application_stats["shortlisted_students"],
            "selected_students": application_stats["selected_students"],
            "unread_notifications": unread_notifications,
        }

        return Response(data)

#placement_officer dashboard 

class PlacementDashboardAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer,
    ]

    def get(self, request):

        total_students = StudentProfile.objects.count()

        total_companies = CompanyProfile.objects.count()

        job_stats = Job.objects.aggregate(
            total_jobs=Count("id"),
            active_jobs=Count("id", filter=Q(is_active=True)),
            inactive_jobs=Count("id", filter=Q(is_active=False)),
        )

        application_stats = JobApplication.objects.aggregate(
            total_applications=Count("id"),
            selected_students=Count(
                "id",
                filter=Q(status="Selected")
            ),
        )

        unread_notifications = Notification.objects.filter(
            recipient=request.user,
            is_read=False
        ).count()

        if total_students > 0:
            placement_percentage = round(
                (application_stats["selected_students"] / total_students) * 100,
                2,
            )
        else:
            placement_percentage = 0

        data = {
            "total_students": total_students,
            "total_companies": total_companies,
            "total_jobs": job_stats["total_jobs"],
            "active_jobs": job_stats["active_jobs"],
            "inactive_jobs": job_stats["inactive_jobs"],
            "total_applications": application_stats["total_applications"],
            "selected_students": application_stats["selected_students"],
            "placement_percentage": placement_percentage,
            "unread_notifications": unread_notifications,
        }

        return Response(data)

#student Report APiView

class StudentReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        students = StudentProfile.objects.select_related(
            "user"
        )

        serializer = StudentProfileSerializer(
            students,
            many=True,
            context={"request": request}
        )

        return Response(serializer.data)


class StudentExcelReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Students Report"

        worksheet.append([
            "Username",
            "Email",
            "Phone",
            "Course",
            "CGPA",
        ])

        students = StudentProfile.objects.select_related(
            "user"
        )

        for student in students:

            worksheet.append([

                student.user.username,

                student.user.email,

                student.phone,

                student.course,

                float(student.cgpa),

            ])

        response = HttpResponse(

            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )

        response["Content-Disposition"] = (

            'attachment; filename="students_report.xlsx"'

        )

        workbook.save(response)

        return response

#studentreport PDF Export

class StudentPDFReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="students_report.pdf"'
        )

        document = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        title = Paragraph(
            "<b>Students Report</b>",
            styles["Title"]
        )

        elements.append(title)

        elements.append(Paragraph("<br/>", styles["Normal"]))

        data = [[
            "Username",
            "Email",
            "Phone",
            "Course",
            "CGPA",
        ]]

        students = StudentProfile.objects.select_related(
            "user"
        )

        for student in students:

            data.append([

                student.user.username,

                student.user.email,

                student.phone,

                student.course,

                str(student.cgpa),

            ])

        table = Table(data)

        table.setStyle(

            TableStyle([

                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563eb")),

                ("TEXTCOLOR", (0,0), (-1,0), colors.white),

                ("GRID", (0,0), (-1,-1), 1, colors.black),

                ("BACKGROUND", (0,1), (-1,-1), colors.beige),

                ("ALIGN", (0,0), (-1,-1), "CENTER"),

                ("BOTTOMPADDING", (0,0), (-1,0), 10),

                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ])

        )

        elements.append(table)

        document.build(elements)

        return response
#company  report APIView

class CompanyReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        companies = CompanyProfile.objects.select_related(
            "user"
        )

        serializer = CompanyProfileSerializer(
            companies,
            many=True,
            context={"request": request}
        )

        return Response(serializer.data)

#company Excel report
class CompanyExcelReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Companies Report"

        worksheet.append([
            "Company Name",
            "Email",
            "Phone",
            "Website",
            "Established Year",
        ])

        companies = CompanyProfile.objects.select_related(
            "user"
        )

        for company in companies:

            worksheet.append([

                company.company_name,

                company.company_email,

                company.phone,

                company.website,

                company.established_year,

            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="companies_report.xlsx"'
        )

        workbook.save(response)

        return response
#company pdf reports
class CompanyPDFReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="companies_report.pdf"'
        )

        document = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Companies Report</b>",
                styles["Title"]
            )
        )

        data = [[
            "Company",
            "Email",
            "Phone",
            "Website",
            "Established",
        ]]

        companies = CompanyProfile.objects.select_related(
            "user"
        )

        for company in companies:

            data.append([

                company.company_name,

                company.company_email,

                company.phone,

                company.website,

                str(company.established_year),

            ])

        table = Table(data)

        table.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563eb")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),1,colors.black),
            ("BACKGROUND",(0,1),(-1,-1),colors.beige),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ]))

        elements.append(table)

        document.build(elements)

        return response
#Jobs ReportAPI VIEw
class JobReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        jobs = Job.objects.select_related(
            "company"
        )

        serializer = JobSerializer(
            jobs,
            many=True,
            context={"request": request}
        )

        return Response(serializer.data)

#JOBS EXCEL REPORT APIVIEW
class JobExcelReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Jobs Report"

        worksheet.append([
            "Job Title",
            "Company",
            "Location",
            "Salary",
            "Minimum CGPA",
            "Deadline",
            "Status",
        ])

        jobs = Job.objects.select_related("company")

        for job in jobs:

            worksheet.append([

                job.job_title,

                job.company.company_name,

                job.location,

                job.salary,

                float(job.minimum_cgpa),

                str(job.last_date),

                "Active" if job.is_active else "Inactive",

            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="jobs_report.xlsx"'
        )

        workbook.save(response)

        return response

#JOBS PDF REPORTS APIVIEW
class JobPDFReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="jobs_report.pdf"'
        )

        document = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Jobs Report</b>",
                styles["Title"]
            )
        )

        data = [[
            "Job",
            "Company",
            "Location",
            "Salary",
            "CGPA",
            "Deadline",
            "Status",
        ]]

        jobs = Job.objects.select_related("company")

        for job in jobs:

            data.append([

                job.job_title,

                job.company.company_name,

                job.location,

                str(job.salary),

                str(job.minimum_cgpa),

                str(job.last_date),

                "Active" if job.is_active else "Inactive",

            ])

        table = Table(data)

        table.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563eb")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),1,colors.black),
            ("BACKGROUND",(0,1),(-1,-1),colors.beige),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ]))

        elements.append(table)

        document.build(elements)

        return response
#Applicant report Apiview
class ApplicationReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        applications = JobApplication.objects.select_related(
            "student__user",
            "job",
            "job__company"
        )

        serializer = JobApplicationSerializer(
            applications,
            many=True,
            context={"request": request}
        )

        return Response(serializer.data)
#applicantreport excel apiview
class ApplicationExcelReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Applications Report"

        sheet.append([
            "Student",
            "Company",
            "Job",
            "Status",
            "Applied Date"
        ])

        applications = JobApplication.objects.select_related(
            "student__user",
            "job",
            "job__company"
        )

        for app in applications:

            sheet.append([

                app.student.user.username,

                app.job.company.company_name,

                app.job.job_title,

                app.status,

                str(app.application_date),

            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="applications_report.xlsx"'
        )

        workbook.save(response)

        return response

#applicant pdf report aPIview

class ApplicationPDFReportAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="applications_report.pdf"'
        )

        document = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Applications Report</b>",
                styles["Title"]
            )
        )

        data = [[
            "Student",
            "Company",
            "Job",
            "Status",
            "Applied Date"
        ]]

        applications = JobApplication.objects.select_related(
            "student__user",
            "job",
            "job__company"
        )

        for app in applications:

            data.append([

                app.student.user.username,

                app.job.company.company_name,

                app.job.job_title,

                app.status,

                str(app.application_date),

            ])

        table = Table(data)

        table.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563eb")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),1,colors.black),
            ("BACKGROUND",(0,1),(-1,-1),colors.beige),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ]))

        elements.append(table)

        document.build(elements)

        return response

#Dashboard Summary API

class DashboardSummaryAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        total_students = StudentProfile.objects.count()

        total_companies = CompanyProfile.objects.count()

        total_jobs = Job.objects.count()

        total_applications = JobApplication.objects.count()

        selected_students = JobApplication.objects.filter(
            status="Selected"
        ).count()

        placement_percentage = 0

        if total_students > 0:

            placement_percentage = round(

                (selected_students / total_students) * 100,

                2

            )

        return Response({

            "total_students": total_students,

            "total_companies": total_companies,

            "total_jobs": total_jobs,

            "total_applications": total_applications,

            "selected_students": selected_students,

            "placement_percentage": placement_percentage,

        })
class ApplicationStatusChartAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        data = JobApplication.objects.values(

            "status"

        ).annotate(

            total=Count("status")

        )

        return Response(data)

class MonthlyJobsAPIView(APIView):

    permission_classes = [
        IsAuthenticated,
        IsPlacementOfficer
    ]

    def get(self, request):

        jobs = Job.objects.annotate(

            month=TruncMonth("created_at")

        ).values(

            "month"

        ).annotate(

            total=Count("id")

        ).order_by("month")

        return Response(jobs)